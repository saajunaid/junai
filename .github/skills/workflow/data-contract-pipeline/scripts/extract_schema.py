"""
Unified Schema Extractor — Reads any structured data source and generates Pydantic models.

Supported formats:
  - JSON (.json) — recursive nesting, list-of-objects, Field(alias=...)
  - Markdown (.md) — tables, bold KV, code blocks, bullet KV, outer wrapper strip
  - CSV/TSV (.csv/.tsv) — column headers, type inference from sample rows
  - XLSX (.xlsx/.xls) — openpyxl, same column logic as CSV
  - YAML (.yaml/.yml) — safe_load, delegates to dict logic
  - Plain text (.txt/.log) — Key: Value, Key = Value, [Section] headers
  - DB table (--format db) — SQLAlchemy inspect + data sampling + embedded format detection

DB discovery & sampling:
  --discover — enumerate all tables/views with columns, PKs, FKs
  --schema <name> — target a specific DB schema (e.g., 'dbo', 'public')
  --sample N — fetch N rows for type inference and embedded format detection
  Embedded formats detected: JSON, XML, YAML, markdown, pipe-delimited in string columns

Schema evolution:
  --save-baseline snap.json — snapshot current schema for later comparison
  --diff snap.json — compare current extraction vs baseline

Usage:
    python extract_schema.py data.json
    python extract_schema.py report.md --class-name NetworkReport
    python extract_schema.py data.csv --output models/ingestion/report.py
    python extract_schema.py data.json --save-baseline baseline.json
    python extract_schema.py data_v2.json --diff baseline.json
    python extract_schema.py --format db --connection-string "mssql+pyodbc://..." --discover
    python extract_schema.py --format db --connection-string "..." --table users --sample 50
    python extract_schema.py --format db --connection-string "..." --table orders --schema dbo

Part of the data-contract-pipeline skill.
"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class ExtractedField:
    name: str
    python_type: str
    source_pattern: str
    example_value: str = ""
    section: str = "Root"


@dataclass
class ExtractedSection:
    name: str
    fields: list[ExtractedField] = field(default_factory=list)
    table_rows_model: str | None = None


@dataclass
class SchemaChange:
    field_name: str
    change_type: str  # added, removed, type_changed
    old_value: str = ""
    new_value: str = ""


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def to_snake_case(name: str) -> str:
    name = name.replace("**", "").strip().rstrip(":")
    s = re.sub(r"[^a-zA-Z0-9]", "_", name)
    s = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", s)
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s.lower()


def to_class_name(name: str) -> str:
    snake = to_snake_case(name)
    return "".join(word.capitalize() for word in snake.split("_"))


def infer_type_from_value(value: Any) -> str:
    if value is None:
        return "Any | None"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, list):
        if not value:
            return "list"
        types = {infer_type_from_value(v) for v in value}
        return f"list[{types.pop()}]" if len(types) == 1 else "list"
    if isinstance(value, dict):
        return "dict"
    v = str(value).strip()
    if not v or v.lower() in ("n/a", "not available", "none", "null", "-"):
        return "str | None"
    if v.lower() in ("true", "false", "yes", "no"):
        return "bool"
    try:
        int(v.replace(",", ""))
        return "int"
    except ValueError:
        pass
    try:
        float(v.replace(",", ""))
        return "float"
    except ValueError:
        pass
    return "str"


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

FORMAT_MAP = {
    ".json": "json",
    ".md": "md",
    ".markdown": "md",
    ".csv": "csv",
    ".tsv": "csv",
    ".xlsx": "xlsx",
    ".xls": "xlsx",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".txt": "text",
    ".log": "text",
}


def detect_format(path: Path | None, explicit: str | None) -> str:
    if explicit and explicit != "auto":
        return explicit
    if path:
        return FORMAT_MAP.get(path.suffix.lower(), "text")
    return "text"


# ---------------------------------------------------------------------------
# JSON extractor
# ---------------------------------------------------------------------------


def extract_json_fields(
    data: Any, section: str = "Root"
) -> list[ExtractedSection]:
    sections: list[ExtractedSection] = []
    if isinstance(data, list):
        data = data[0] if data else {}
    if not isinstance(data, dict):
        return sections

    root = ExtractedSection(name=section)
    for key, value in data.items():
        snake = to_snake_case(key)
        if not snake or not snake.replace("_", "").isalnum():
            continue

        if isinstance(value, dict) and value:
            child_class = to_class_name(key)
            child_sections = extract_json_fields(value, section=key)
            sections.extend(child_sections)
            root.fields.append(
                ExtractedField(
                    name=snake,
                    python_type=child_class,
                    source_pattern="json_nested",
                    example_value=f"{len(value)} keys",
                    section=section,
                )
            )
        elif isinstance(value, list) and value and isinstance(value[0], dict):
            row_class = to_class_name(key) + "Item"
            row_sections = extract_json_fields(value[0], section=f"{key} (Item)")
            for rs in row_sections:
                if rs.name == f"{key} (Item)":
                    rs.name = f"{key} (Row Model)"
            sections.extend(row_sections)
            root.fields.append(
                ExtractedField(
                    name=snake,
                    python_type=f"list[{row_class}]",
                    source_pattern="json_list",
                    example_value=f"{len(value)} items",
                    section=section,
                )
            )
        else:
            example = str(value)[:80] if value is not None else "null"
            root.fields.append(
                ExtractedField(
                    name=snake,
                    python_type=infer_type_from_value(value),
                    source_pattern="json_key",
                    example_value=example,
                    section=section,
                )
            )

    if root.fields:
        sections.insert(0, root)
    return sections


# ---------------------------------------------------------------------------
# Markdown extractor
# ---------------------------------------------------------------------------


def strip_outer_table_wrapper(text: str) -> str:
    """Strip single-cell table wrapper (| col | / | --- | header) from LLM outputs."""
    lines = text.strip().split("\n")
    if len(lines) < 3:
        return text
    header = lines[0].strip()
    if not re.match(r"^\|[^|]+\|\s*$", header):
        return text
    separator = lines[1].strip()
    if not re.match(r"^\|[-\s:]+\|\s*$", separator):
        return text
    content_lines = lines[2:]
    if content_lines:
        first = content_lines[0].strip()
        if first.startswith("|"):
            content_lines[0] = first[1:].lstrip()
    return "\n".join(content_lines)


def parse_markdown_tables(text: str) -> list[dict]:
    tables: list[dict] = []
    lines = text.split("\n")
    i = 0
    while i < len(lines) - 1:
        line = lines[i].strip()
        # Separator regex: dash MUST be first in char class to avoid range
        if re.match(r"^\|[-\s:]+\|[-\s:|]+\|$", line.replace(" ", "")):
            if i > 0 and "|" in lines[i - 1]:
                header_row = lines[i - 1].strip()
                data_rows: list[str] = []
                j = i + 1
                while j < len(lines) and "|" in lines[j] and lines[j].strip():
                    data_rows.append(lines[j].strip())
                    j += 1
                header_cells = [
                    c.strip()
                    for c in header_row.strip().strip("|").split("|")
                    if c.strip()
                ]
                tables.append(
                    {
                        "header_row": header_row,
                        "data_rows": data_rows,
                        "column_count": len(header_cells),
                    }
                )
                i = j
                continue
        i += 1
    return tables


def extract_bold_kv(text: str) -> list[ExtractedField]:
    fields: list[ExtractedField] = []
    for m in re.finditer(
        r"\*\*(.+?)\*\*\s*:?\s*(.+?)(?:\s{2,}|$)", text, re.MULTILINE
    ):
        key = m.group(1).strip().rstrip(":")
        value = m.group(2).strip()
        if value.startswith("**") or value.startswith("#"):
            continue
        snake = to_snake_case(key)
        if not snake or not snake.replace("_", "").isalnum():
            continue
        fields.append(
            ExtractedField(
                name=snake,
                python_type=infer_type_from_value(value),
                source_pattern="bold_kv",
                example_value=value[:80],
            )
        )
    return fields


def extract_code_block_kv(block: str) -> list[ExtractedField]:
    fields: list[ExtractedField] = []
    seen: set[str] = set()
    for line in block.split("\n"):
        line = line.strip()
        if not line or line.startswith("\u2501") or line.startswith("---"):
            continue
        for seg in line.split("|"):
            seg = seg.strip()
            m = re.match(r"^([A-Za-z][A-Za-z0-9 _/#()-]+?)\s*:\s*(.+)$", seg)
            if m:
                snake = to_snake_case(m.group(1).strip())
                if (
                    snake
                    and snake.replace("_", "").isalnum()
                    and snake not in seen
                ):
                    seen.add(snake)
                    fields.append(
                        ExtractedField(
                            name=snake,
                            python_type=infer_type_from_value(
                                m.group(2).strip()
                            ),
                            source_pattern="code_block_kv",
                            example_value=m.group(2).strip()[:80],
                        )
                    )
    return fields


def extract_bullet_kv(
    text: str, already: set[str] | None = None
) -> list[ExtractedField]:
    """Extract non-bold bullet KV pairs. Bold bullets are handled by bold_kv."""
    if already is None:
        already = set()
    fields: list[ExtractedField] = []
    seen: set[str] = set()
    for m in re.finditer(
        r"^[-*]\s+(?!\*\*)([^*:]+?)\s*:\s*(.+)$", text, re.MULTILINE
    ):
        snake = to_snake_case(m.group(1).strip())
        if (
            snake
            and snake.replace("_", "").isalnum()
            and snake not in seen
            and snake not in already
        ):
            seen.add(snake)
            fields.append(
                ExtractedField(
                    name=snake,
                    python_type=infer_type_from_value(m.group(2).strip()),
                    source_pattern="bullet_kv",
                    example_value=m.group(2).strip()[:80],
                )
            )
    return fields


def extract_md_sections(text: str) -> list[tuple[str, str]]:
    sections: list[tuple[str, str]] = []
    matches = list(re.finditer(r"^#{1,3}\s+(.+)$", text, re.MULTILINE))
    if not matches:
        return [("Root", text)]
    if matches[0].start() > 0:
        sections.append(("Root", text[: matches[0].start()]))
    for i, m in enumerate(matches):
        heading = re.sub(r"^\d+\.\s*", "", m.group(1).strip())
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        sections.append((heading, text[m.end() : end].strip()))
    return sections


def extract_markdown(text: str) -> list[ExtractedSection]:
    text = strip_outer_table_wrapper(text)
    result: list[ExtractedSection] = []
    for heading, content in extract_md_sections(text):
        section = ExtractedSection(name=heading)

        # Markdown tables
        for table in parse_markdown_tables(content):
            if table["column_count"] == 2:
                for row in table["data_rows"]:
                    cells = [
                        c.strip().replace("**", "")
                        for c in row.strip().strip("|").split("|")
                    ]
                    if len(cells) >= 2:
                        snake = to_snake_case(cells[0])
                        if snake and snake.replace("_", "").isalnum():
                            section.fields.append(
                                ExtractedField(
                                    name=snake,
                                    python_type=infer_type_from_value(
                                        cells[1].strip()
                                    ),
                                    source_pattern="md_table_kv",
                                    example_value=cells[1].strip()[:80],
                                )
                            )
            elif table["column_count"] >= 3:
                headers = [
                    c.strip().replace("**", "")
                    for c in table["header_row"].strip().strip("|").split("|")
                    if c.strip()
                ]
                first_vals = (
                    [
                        c.strip().replace("**", "")
                        for c in table["data_rows"][0]
                        .strip()
                        .strip("|")
                        .split("|")
                        if c.strip()
                    ]
                    if table["data_rows"]
                    else []
                )
                row_fields = []
                for idx, h in enumerate(headers):
                    snake = to_snake_case(h)
                    if snake and snake.replace("_", "").isalnum():
                        val = first_vals[idx] if idx < len(first_vals) else ""
                        row_fields.append(
                            ExtractedField(
                                name=snake,
                                python_type=infer_type_from_value(val),
                                source_pattern="md_table_multi",
                                example_value=val[:80],
                            )
                        )
                row_class = to_class_name(heading) + "Row"
                result.append(
                    ExtractedSection(
                        name=f"{heading} (Row Model)", fields=row_fields
                    )
                )
                section.fields.append(
                    ExtractedField(
                        name=to_snake_case(heading) + "_rows",
                        python_type=f"list[{row_class}]",
                        source_pattern="md_table_multi",
                        example_value=f"{len(table['data_rows'])} rows",
                    )
                )

        # Code blocks
        for block in re.findall(r"```[^\n]*\n(.*?)```", content, re.DOTALL):
            section.fields.extend(extract_code_block_kv(block))

        # Bold + bullet KV (after removing tables and code blocks)
        clean = re.sub(r"```.*?```", "", content, flags=re.DOTALL)
        clean = re.sub(r"^\|.+\|$", "", clean, flags=re.MULTILINE)
        bold_fields = extract_bold_kv(clean)
        section.fields.extend(bold_fields)
        bold_names = {f.name for f in bold_fields}
        section.fields.extend(extract_bullet_kv(clean, already=bold_names))

        if section.fields:
            if heading == "Root":
                result.insert(0, section)
            else:
                result.append(section)
    return result


# ---------------------------------------------------------------------------
# CSV extractor
# ---------------------------------------------------------------------------


def extract_csv(
    text: str, delimiter: str | None = None
) -> list[ExtractedSection]:
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(text[:4096])
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if len(rows) < 2:
        return []

    headers = rows[0]
    data_rows = rows[1:101]  # sample up to 100 rows for type inference

    section = ExtractedSection(name="Root")
    for col_idx, header in enumerate(headers):
        snake = to_snake_case(header)
        if not snake or not snake.replace("_", "").isalnum():
            snake = f"col_{col_idx}"

        sample_values = [
            r[col_idx]
            for r in data_rows
            if col_idx < len(r) and r[col_idx].strip()
        ]
        if sample_values:
            types = {infer_type_from_value(v) for v in sample_values[:20]}
            if len(types) == 1:
                py_type = types.pop()
            elif types <= {"int", "float"}:
                py_type = "float"
            elif "str | None" in types:
                py_type = "str | None"
            else:
                py_type = "str"
            example = sample_values[0][:80]
        else:
            py_type = "str | None"
            example = ""

        section.fields.append(
            ExtractedField(
                name=snake,
                python_type=py_type,
                source_pattern="csv_col",
                example_value=example,
            )
        )
    return [section]


# ---------------------------------------------------------------------------
# XLSX extractor
# ---------------------------------------------------------------------------


def extract_xlsx(path: Path) -> list[ExtractedSection]:
    try:
        import openpyxl
    except ImportError:
        print(
            "Error: openpyxl required for xlsx. Install: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_sections: list[ExtractedSection] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [
            str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])
        ]
        data_rows = rows[1:101]

        section = ExtractedSection(
            name=sheet_name if len(wb.sheetnames) > 1 else "Root"
        )
        for col_idx, header in enumerate(headers):
            snake = to_snake_case(header)
            if not snake or not snake.replace("_", "").isalnum():
                snake = f"col_{col_idx}"

            samples = [
                r[col_idx]
                for r in data_rows
                if col_idx < len(r) and r[col_idx] is not None
            ]
            if samples:
                py_type = infer_type_from_value(samples[0])
                example = str(samples[0])[:80]
            else:
                py_type = "str | None"
                example = ""

            section.fields.append(
                ExtractedField(
                    name=snake,
                    python_type=py_type,
                    source_pattern="xlsx_col",
                    example_value=example,
                )
            )
        all_sections.append(section)

    wb.close()
    return all_sections


# ---------------------------------------------------------------------------
# YAML extractor
# ---------------------------------------------------------------------------


def extract_yaml(text: str) -> list[ExtractedSection]:
    try:
        import yaml
    except ImportError:
        print(
            "Error: PyYAML required. Install: pip install pyyaml",
            file=sys.stderr,
        )
        sys.exit(1)
    data = yaml.safe_load(text)
    return extract_json_fields(data)


# ---------------------------------------------------------------------------
# Plain text extractor
# ---------------------------------------------------------------------------


def extract_text(text: str) -> list[ExtractedSection]:
    sections: list[ExtractedSection] = []
    current_section = "Root"
    current_fields: list[ExtractedField] = []

    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue

        # INI-style section header
        sec_match = re.match(r"^\[(.+)\]$", line)
        if sec_match:
            if current_fields:
                sections.append(
                    ExtractedSection(
                        name=current_section, fields=current_fields
                    )
                )
            current_section = sec_match.group(1).strip()
            current_fields = []
            continue

        # Key: Value or Key = Value
        kv_match = re.match(
            r"^([A-Za-z][A-Za-z0-9 _.-]+?)\s*[:=]\s*(.+)$", line
        )
        if kv_match:
            snake = to_snake_case(kv_match.group(1).strip())
            value = kv_match.group(2).strip()
            if snake and snake.replace("_", "").isalnum():
                current_fields.append(
                    ExtractedField(
                        name=snake,
                        python_type=infer_type_from_value(value),
                        source_pattern="text_kv",
                        example_value=value[:80],
                    )
                )

    if current_fields:
        sections.append(
            ExtractedSection(name=current_section, fields=current_fields)
        )
    return sections


# ---------------------------------------------------------------------------
# DB table extractor
# ---------------------------------------------------------------------------

SQL_TYPE_MAP = {
    # SQL Server
    "INTEGER": "int",
    "BIGINT": "int",
    "SMALLINT": "int",
    "TINYINT": "int",
    "FLOAT": "float",
    "REAL": "float",
    "DECIMAL": "float",
    "NUMERIC": "float",
    "MONEY": "float",
    "SMALLMONEY": "float",
    "BIT": "bool",
    "VARCHAR": "str",
    "NVARCHAR": "str",
    "CHAR": "str",
    "NCHAR": "str",
    "TEXT": "str",
    "NTEXT": "str",
    "DATE": "str",
    "DATETIME": "str",
    "DATETIME2": "str",
    "SMALLDATETIME": "str",
    "TIME": "str",
    "DATETIMEOFFSET": "str",
    "UNIQUEIDENTIFIER": "str",
    "VARBINARY": "bytes",
    "BINARY": "bytes",
    "IMAGE": "bytes",
    # PostgreSQL
    "SERIAL": "int",
    "BIGSERIAL": "int",
    "SMALLSERIAL": "int",
    "BOOLEAN": "bool",
    "DOUBLE_PRECISION": "float",
    "JSONB": "dict",
    "JSON": "dict",
    "UUID": "str",
    "BYTEA": "bytes",
    "TIMESTAMPTZ": "str",
    "TIMESTAMP": "str",
    "INTERVAL": "str",
    "CITEXT": "str",
    "INET": "str",
    "MACADDR": "str",
    "ARRAY": "list",
    "HSTORE": "dict",
    # MySQL
    "TINYTEXT": "str",
    "MEDIUMTEXT": "str",
    "LONGTEXT": "str",
    "MEDIUMINT": "int",
    "ENUM": "str",
    "SET": "str",
    "BLOB": "bytes",
    "MEDIUMBLOB": "bytes",
    "LONGBLOB": "bytes",
    "TINYBLOB": "bytes",
    # SQLite
    "REAL": "float",
}


def _detect_embedded_format(value: str) -> str | None:
    """Detect structured data embedded in a string column value."""
    v = value.strip()
    if not v:
        return None
    if (v.startswith("{") and v.endswith("}")) or (
        v.startswith("[") and v.endswith("]")
    ):
        try:
            json.loads(v)
            return "json"
        except (json.JSONDecodeError, ValueError):
            pass
    if v.startswith("<?xml") or (v.startswith("<") and ">" in v):
        return "xml"
    if v.startswith("---\n") or re.match(r"^[a-zA-Z_]+\s*:", v):
        try:
            import yaml

            yaml.safe_load(v)
            return "yaml"
        except Exception:
            pass
    if "|" in v and v.count("|") >= 3:
        return "pipe-delimited"
    if v.startswith("#") or v.startswith("**") or re.search(r"\[.+\]\(.+\)", v):
        return "markdown"
    return None


def _sample_rows(
    engine: Any, table_name: str, schema: str | None, limit: int = 50
) -> list[dict]:
    """Fetch sample rows from a table for type inference and format detection."""
    from sqlalchemy import text

    qualified = f'"{schema}"."{table_name}"' if schema else f'"{table_name}"'
    query = text(f"SELECT * FROM {qualified} ORDER BY 1 LIMIT :lim")
    try:
        with engine.connect() as conn:
            result = conn.execute(query, {"lim": limit})
            columns = list(result.keys())
            return [dict(zip(columns, row)) for row in result.fetchall()]
    except Exception:
        # MSSQL uses TOP instead of LIMIT
        query_mssql = text(
            f"SELECT TOP(:lim) * FROM {qualified} ORDER BY 1"
        )
        try:
            with engine.connect() as conn:
                result = conn.execute(query_mssql, {"lim": limit})
                columns = list(result.keys())
                return [dict(zip(columns, row)) for row in result.fetchall()]
        except Exception:
            return []


def discover_db_tables(
    connection_string: str, schema: str | None = None
) -> list[dict]:
    """Discover all tables and views in a database schema.

    Returns a list of dicts with keys: name, type (table/view), column_count,
    row_count_estimate, columns (list of column info dicts).
    """
    try:
        from sqlalchemy import create_engine, text
        from sqlalchemy import inspect as sa_inspect
    except ImportError:
        print(
            "Error: sqlalchemy required. Install: pip install sqlalchemy",
            file=sys.stderr,
        )
        sys.exit(1)

    engine = create_engine(connection_string)
    inspector = sa_inspect(engine)

    tables = []
    # Get tables
    for table_name in inspector.get_table_names(schema=schema):
        columns = inspector.get_columns(table_name, schema=schema)
        pk = inspector.get_pk_constraint(table_name, schema=schema)
        fks = inspector.get_foreign_keys(table_name, schema=schema)
        tables.append({
            "name": table_name,
            "schema": schema,
            "type": "table",
            "column_count": len(columns),
            "columns": [
                {
                    "name": c["name"],
                    "sql_type": type(c["type"]).__name__.upper(),
                    "nullable": c.get("nullable", True),
                    "is_pk": c["name"] in (pk.get("constrained_columns", []) if pk else []),
                }
                for c in columns
            ],
            "primary_key": pk.get("constrained_columns", []) if pk else [],
            "foreign_keys": [
                {
                    "columns": fk["constrained_columns"],
                    "referred_table": fk["referred_table"],
                    "referred_columns": fk["referred_columns"],
                }
                for fk in fks
            ],
        })

    # Get views
    for view_name in inspector.get_view_names(schema=schema):
        columns = inspector.get_columns(view_name, schema=schema)
        tables.append({
            "name": view_name,
            "schema": schema,
            "type": "view",
            "column_count": len(columns),
            "columns": [
                {
                    "name": c["name"],
                    "sql_type": type(c["type"]).__name__.upper(),
                    "nullable": c.get("nullable", True),
                    "is_pk": False,
                }
                for c in columns
            ],
            "primary_key": [],
            "foreign_keys": [],
        })

    engine.dispose()
    return tables


def extract_db_table(
    connection_string: str,
    table_name: str,
    schema: str | None = None,
    sample: int = 0,
) -> list[ExtractedSection]:
    """Extract schema from a DB table with optional data sampling.

    Args:
        connection_string: SQLAlchemy connection string
        table_name: Table or view name
        schema: Optional DB schema (e.g., 'dbo', 'public')
        sample: Number of rows to sample for type inference and format detection.
                0 = metadata only (original behavior).
    """
    try:
        from sqlalchemy import create_engine
        from sqlalchemy import inspect as sa_inspect
    except ImportError:
        print(
            "Error: sqlalchemy required. Install: pip install sqlalchemy",
            file=sys.stderr,
        )
        sys.exit(1)

    engine = create_engine(connection_string)
    inspector = sa_inspect(engine)

    columns = inspector.get_columns(table_name, schema=schema)
    pk = inspector.get_pk_constraint(table_name, schema=schema)
    pk_cols = pk.get("constrained_columns", []) if pk else []
    fks = inspector.get_foreign_keys(table_name, schema=schema)

    # Sample rows if requested
    sampled_rows: list[dict] = []
    if sample > 0:
        sampled_rows = _sample_rows(engine, table_name, schema, limit=sample)

    section = ExtractedSection(name="Root")
    embedded_sections: list[ExtractedSection] = []

    for col in columns:
        snake = to_snake_case(col["name"])
        sql_type = type(col["type"]).__name__.upper()
        py_type = SQL_TYPE_MAP.get(sql_type, "Any")
        if col.get("nullable", True):
            py_type = f"{py_type} | None"

        # Build example from sampled data
        example = f"SQL: {sql_type}"
        embedded_fmt = None
        if sampled_rows:
            sample_vals = [
                row.get(col["name"])
                for row in sampled_rows
                if row.get(col["name"]) is not None
            ]
            if sample_vals:
                example = str(sample_vals[0])[:80]
                # Check for embedded structured data in string columns
                if sql_type in (
                    "VARCHAR", "NVARCHAR", "TEXT", "NTEXT",
                    "MEDIUMTEXT", "LONGTEXT", "CITEXT",
                ):
                    for sv in sample_vals[:10]:
                        fmt = _detect_embedded_format(str(sv))
                        if fmt:
                            embedded_fmt = fmt
                            break

        source_pattern = "db_column"
        if col["name"] in pk_cols:
            source_pattern = "db_column_pk"
        if embedded_fmt:
            source_pattern = f"db_column_embedded_{embedded_fmt}"
            example = f"[embedded {embedded_fmt}] {example}"
            # Parse the embedded content to extract nested schema
            if embedded_fmt == "json" and sampled_rows:
                for sv in sample_vals[:1]:
                    try:
                        parsed = json.loads(str(sv))
                        nested = extract_json_fields(
                            parsed, section=f"{col['name']} (Embedded JSON)"
                        )
                        embedded_sections.extend(nested)
                    except (json.JSONDecodeError, ValueError):
                        pass

        section.fields.append(
            ExtractedField(
                name=snake,
                python_type=py_type,
                source_pattern=source_pattern,
                example_value=example,
            )
        )

    # Add FK annotations as comments in the section
    for fk in fks:
        fk_desc = (
            f"FK: {','.join(fk['constrained_columns'])} -> "
            f"{fk['referred_table']}({','.join(fk['referred_columns'])})"
        )
        section.fields.append(
            ExtractedField(
                name=f"_fk_{fk['referred_table']}",
                python_type="# relationship",
                source_pattern="db_foreign_key",
                example_value=fk_desc,
            )
        )

    engine.dispose()
    result = []
    if section.fields:
        result.append(section)
    result.extend(embedded_sections)
    return result


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------


def deduplicate_fields(
    fields: list[ExtractedField],
) -> list[ExtractedField]:
    seen: set[str] = set()
    result: list[ExtractedField] = []
    for f in fields:
        if f.name not in seen:
            seen.add(f.name)
            result.append(f)
    return result


# ---------------------------------------------------------------------------
# Model code generation (shared across all formats)
# ---------------------------------------------------------------------------


def generate_model_source(
    sections: list[ExtractedSection], root_class_name: str = "GeneratedPayload"
) -> str:
    lines = [
        '"""',
        "Auto-generated ingestion model.",
        "Generated by: extract_schema.py",
        "",
        "Review and adjust:",
        "  - Field names and types",
        "  - Which sections become nested models vs flat fields",
        "  - Optional fields and defaults",
        '"""',
        "",
        "from __future__ import annotations",
        "",
        "from pydantic import BaseModel, ConfigDict, Field",
        "",
    ]

    row_models: list[ExtractedSection] = []
    content_sections: list[ExtractedSection] = []
    for s in sections:
        if s.name.endswith("(Row Model)"):
            row_models.append(s)
        else:
            content_sections.append(s)

    # Nested/row models first
    for rm in row_models:
        cname = to_class_name(rm.name.replace(" (Row Model)", "")) + "Row"
        if "(Item)" in rm.name:
            cname = (
                to_class_name(
                    rm.name.replace(" (Item)", "").replace(
                        " (Row Model)", ""
                    )
                )
                + "Item"
            )
        rm_fields = deduplicate_fields(rm.fields)
        lines += [
            "",
            f"class {cname}(BaseModel):",
            f'    """Row from {rm.name.replace(" (Row Model)", "")}."""',
            "",
        ]
        if not rm_fields:
            lines.append("    pass")
        else:
            for f in rm_fields:
                comment = (
                    f"  # e.g., {f.example_value}" if f.example_value else ""
                )
                lines.append(f"    {f.name}: {f.python_type}{comment}")
        lines.append("")

    # Root model
    all_fields: list[ExtractedField] = []
    section_markers: list[tuple[int, str]] = []
    for cs in content_sections:
        cs_fields = deduplicate_fields(cs.fields)
        if cs_fields:
            section_markers.append((len(all_fields), cs.name))
            all_fields.extend(cs_fields)
    all_fields = deduplicate_fields(all_fields)

    lines += [
        "",
        f"class {root_class_name}(BaseModel):",
        f'    """1:1 representation of source payload."""',
        "",
        '    model_config = ConfigDict(extra="allow")',
        "",
    ]

    if not all_fields:
        lines.append("    pass")
    else:
        marker_idx = 0
        for i, f in enumerate(all_fields):
            if marker_idx < len(section_markers):
                fidx, sname = section_markers[marker_idx]
                if i >= fidx:
                    lines.append(f"    # --- {sname} ---")
                    marker_idx += 1
            comment = (
                f"  # e.g., {f.example_value}" if f.example_value else ""
            )
            lines.append(f"    {f.name}: {f.python_type}{comment}")

    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Schema evolution
# ---------------------------------------------------------------------------


def sections_to_schema_dict(
    sections: list[ExtractedSection], source: str, fmt: str
) -> dict:
    fields = []
    for s in sections:
        for f in s.fields:
            fields.append(
                {
                    "name": f.name,
                    "type": f.python_type,
                    "source_pattern": f.source_pattern,
                    "section": s.name,
                }
            )
    return {
        "extracted_at": datetime.now(timezone.utc).isoformat(),
        "source_file": source,
        "format": fmt,
        "field_count": len(fields),
        "fields": fields,
    }


def save_baseline(
    sections: list[ExtractedSection], path: Path, source: str, fmt: str
) -> None:
    schema = sections_to_schema_dict(sections, source, fmt)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(schema, indent=2), encoding="utf-8")
    print(
        f"Baseline saved: {path} ({schema['field_count']} fields)",
        file=sys.stderr,
    )


def diff_baseline(
    sections: list[ExtractedSection], baseline_path: Path
) -> list[SchemaChange]:
    baseline = json.loads(baseline_path.read_text(encoding="utf-8"))
    old_fields = {f["name"]: f["type"] for f in baseline["fields"]}
    new_fields: dict[str, str] = {}
    for s in sections:
        for f in s.fields:
            if f.name not in new_fields:
                new_fields[f.name] = f.python_type

    changes: list[SchemaChange] = []
    for name in sorted(set(old_fields) | set(new_fields)):
        if name not in old_fields:
            changes.append(
                SchemaChange(name, "added", new_value=new_fields[name])
            )
        elif name not in new_fields:
            changes.append(
                SchemaChange(name, "removed", old_value=old_fields[name])
            )
        elif old_fields[name] != new_fields[name]:
            changes.append(
                SchemaChange(
                    name,
                    "type_changed",
                    old_value=old_fields[name],
                    new_value=new_fields[name],
                )
            )
    return changes


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(
        description="Unified schema extractor — any format to Pydantic"
    )
    parser.add_argument(
        "source",
        type=Path,
        nargs="?",
        default=None,
        help="Path to source data file",
    )
    parser.add_argument(
        "--format",
        dest="fmt",
        default="auto",
        choices=["auto", "json", "md", "csv", "xlsx", "yaml", "db", "text"],
        help="Source format (default: auto-detect from extension)",
    )
    parser.add_argument(
        "--class-name",
        default="GeneratedPayload",
        help="Root class name",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output .py file path",
    )
    parser.add_argument(
        "--show-sections",
        action="store_true",
        help="Print extracted sections instead of code",
    )
    parser.add_argument(
        "--save-baseline",
        type=Path,
        default=None,
        help="Save schema snapshot to JSON",
    )
    parser.add_argument(
        "--diff",
        type=Path,
        default=None,
        help="Diff current schema against baseline JSON",
    )
    parser.add_argument(
        "--connection-string",
        default=None,
        help="DB connection string (for --format db)",
    )
    parser.add_argument(
        "--table",
        default=None,
        help="DB table name (for --format db)",
    )
    parser.add_argument(
        "--schema",
        default=None,
        help="DB schema name, e.g. 'dbo', 'public' (for --format db)",
    )
    parser.add_argument(
        "--discover",
        action="store_true",
        help="Discover all tables/views in the database (for --format db)",
    )
    parser.add_argument(
        "--sample",
        type=int,
        default=0,
        help="Number of rows to sample for type inference and embedded format detection (for --format db, default: 0 = metadata only)",
    )
    args = parser.parse_args()

    fmt = detect_format(args.source, args.fmt)

    if fmt == "db":
        if not args.connection_string:
            print(
                "Error: --connection-string required for db format",
                file=sys.stderr,
            )
            sys.exit(1)

        # Discovery mode: list all tables and their schemas
        if args.discover:
            tables = discover_db_tables(args.connection_string, schema=args.schema)
            print(f"Discovered {len(tables)} tables/views:\n")
            for t in tables:
                pk_str = f"  PK: {', '.join(t['primary_key'])}" if t['primary_key'] else ""
                fk_str = f"  FKs: {len(t['foreign_keys'])}" if t['foreign_keys'] else ""
                print(f"  [{t['type'].upper()}] {t['name']} ({t['column_count']} columns){pk_str}{fk_str}")
                for col in t['columns']:
                    pk_marker = " [PK]" if col['is_pk'] else ""
                    null_marker = " NULL" if col['nullable'] else " NOT NULL"
                    print(f"    {col['name']}: {col['sql_type']}{null_marker}{pk_marker}")
                if t['foreign_keys']:
                    for fk in t['foreign_keys']:
                        print(f"    FK: {','.join(fk['columns'])} -> {fk['referred_table']}({','.join(fk['referred_columns'])})")
                print()
            sys.exit(0)

        if not args.table:
            print(
                "Error: --table required for db format (or use --discover to list tables)",
                file=sys.stderr,
            )
            sys.exit(1)
        sections = extract_db_table(
            args.connection_string, args.table,
            schema=args.schema, sample=args.sample,
        )
        source_label = f"{args.table} (db)"
    else:
        if not args.source or not args.source.exists():
            print(
                f"Error: File not found: {args.source}", file=sys.stderr
            )
            sys.exit(1)
        source_label = str(args.source)

        if fmt == "xlsx":
            sections = extract_xlsx(args.source)
        else:
            text = args.source.read_text(encoding="utf-8")
            extractors = {
                "json": lambda t: extract_json_fields(json.loads(t)),
                "md": extract_markdown,
                "csv": extract_csv,
                "yaml": extract_yaml,
                "text": extract_text,
            }
            sections = extractors[fmt](text)

    if args.save_baseline:
        save_baseline(sections, args.save_baseline, source_label, fmt)

    if args.diff:
        if not args.diff.exists():
            print(
                f"Error: Baseline not found: {args.diff}", file=sys.stderr
            )
            sys.exit(1)
        changes = diff_baseline(sections, args.diff)
        if not changes:
            print("No schema changes detected.")
        else:
            print(f"{len(changes)} schema change(s) detected:\n")
            for c in changes:
                if c.change_type == "added":
                    print(f"  + ADDED   {c.field_name}: {c.new_value}")
                elif c.change_type == "removed":
                    print(f"  - REMOVED {c.field_name}: {c.old_value}")
                else:
                    print(
                        f"  ~ CHANGED {c.field_name}: {c.old_value} -> {c.new_value}"
                    )
        sys.exit(0)

    if args.show_sections:
        print(f"Extracted {len(sections)} sections:\n")
        for s in sections:
            print(f"  [{s.name}] -- {len(s.fields)} fields")
            for f in s.fields:
                print(
                    f"    {f.name}: {f.python_type} ({f.source_pattern}) = {f.example_value!r}"
                )
            print()
        sys.exit(0)

    source_code = generate_model_source(
        sections, root_class_name=args.class_name
    )
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(source_code, encoding="utf-8")
        print(f"Written to {args.output}")
    else:
        print(source_code)


if __name__ == "__main__":
    main()
